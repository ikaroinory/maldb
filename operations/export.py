import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from path import get_export_path

with open('./sql/malware_info.sql', 'r') as file:
    malware_info_sql = file.read()
with open('./sql/total_count.sql', 'r') as file:
    total_count_sql = file.read()
with open('./sql/category1_statistic.sql', 'r') as file:
    category1_statistic_sql = file.read()
with open('./sql/category2_statistic.sql', 'r') as file:
    category2_statistic_sql = file.read()
with open('./sql/category3_statistic.sql', 'r') as file:
    category3_statistic_sql = file.read()
with open('./sql/name_statistic.sql', 'r') as file:
    name_statistic_sql = file.read()


def export(db_path: str, path: str | None) -> None:
    if path is None:
        path = get_export_path()

    excel_path = Path(path) / 'export_info.maldb.xlsx'

    with sqlite3.connect(db_path) as conn:
        malware_info_df = pd.read_sql_query(malware_info_sql, conn)
        total_count_df = pd.read_sql_query(total_count_sql, conn)
        category1_statistic_df = pd.read_sql_query(category1_statistic_sql, conn)
        category2_statistic_df = pd.read_sql_query(category2_statistic_sql, conn)
        category3_statistic_df = pd.read_sql_query(category3_statistic_sql, conn)
        name_statistic_df = pd.read_sql_query(name_statistic_sql, conn)

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        malware_info_df.to_excel(writer, sheet_name='Malware Samples Information', index=False)
        total_count_df.to_excel(writer, sheet_name='Statistic', index=False, startrow=1, startcol=1)
        category1_statistic_df.to_excel(writer, sheet_name='Statistic', index=False, startrow=1, startcol=3)
        category1_statistic_df.to_excel(writer, sheet_name='Statistic', index=False, startrow=1, startcol=3)
        category2_statistic_df.to_excel(writer, sheet_name='Statistic', index=False, startrow=1, startcol=7)
        category3_statistic_df.to_excel(writer, sheet_name='Statistic', index=False, startrow=1, startcol=11)
        name_statistic_df.to_excel(writer, sheet_name='Statistic', index=False, startrow=1, startcol=15)

        courier_new_font = Font(name='Courier New', size=12)
        courier_new_bold_font = Font(name='Courier New', size=12, bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_side = Side(border_style="thin", color="000000")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        wb = writer.book
        ws = wb['Malware Samples Information']

        ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'

        column_index = 1
        for column in ws.columns:
            max_len = 0
            column_letter = column[0].column_letter
            for cell in column:
                cell.font = courier_new_font
                cell.border = border
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))

                if column_index in [1, 2, 3, 4, 5, 7, 8, 9, 10, 11, 13, 14, 15, 16]:
                    cell.alignment = center_alignment
            ws.column_dimensions[column_letter].width = max_len * 1.5
            column_index += 1

        for cell in ws[1]:
            cell.font = courier_new_bold_font
            cell.alignment = center_alignment

        ws = wb['Statistic']
        column_index = 1
        for column in ws.columns:
            max_len = 0
            column_letter = column[0].column_letter
            for cell in column:
                cell.font = courier_new_font
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = max_len * 1.5
            column_index += 1

        ws['B3'].alignment = center_alignment

        for column in ws.columns:
            for cell in column:
                if cell.value:
                    cell.border = border
        for cell in ws[2]:
            cell.font = courier_new_bold_font
            cell.alignment = center_alignment

        for cell in ws['F']:
            cell.number_format = '0.00%'
        for cell in ws['J']:
            cell.number_format = '0.00%'
        for cell in ws['N']:
            cell.number_format = '0.00%'

    print(f'Exported {excel_path} successfully.')
